from datetime import datetime
import openpyxl
import numpy as np 

class check_in_out():
    def __init__(self):
        self.workbook = openpyxl.load_workbook('data\\data2.xlsx')
        self.status = "Default status"
        self.time_check = "Default time_check"

    def save_data(self, sheet_index):
        now = datetime.now()
        dt_string = now.strftime("%d/%m/%Y %H:%M:%S")
        date = dt_string.split(" ")[0]
        time = dt_string.split(" ")[1]
        
        self.workbook._active_sheet_index = sheet_index
        sheet = self.workbook.active
        
        next_row = sheet.max_row + 1
        day = sheet.cell(row=next_row-1, column=1)
        in_ = sheet.cell(row=next_row-1, column=2)
        out_ = sheet.cell(row=next_row-1, column=3)
        
        if str(day.value) == "วันที่" or str(day.value) != str(date):
            sheet.cell(row=next_row, column=1, value=date)
            sheet.cell(row=next_row, column=2, value=time) 
            self.status = "Checked in"
            self.time_check = str(time)
        elif str(day.value) == str(date) and in_.value == None:
            sheet.cell(row=next_row-1, column=1, value=date)
            sheet.cell(row=next_row-1, column=2, value=time) 
            self.status = "Checked in"
            self.time_check = str(time)
        elif str(day.value) == str(date) and out_.value == None:
            sheet.cell(row=next_row-1, column=1, value=date)
            sheet.cell(row=next_row-1, column=3, value=time)
            self.status = "Checked out"
            self.time_check = str(time)
        else:
            self.status = ""
            self.time_check = ""
        self.workbook.save('data\\data2.xlsx')  

        return np.array([self.status, self.time_check])

